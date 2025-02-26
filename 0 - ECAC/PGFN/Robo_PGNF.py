import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.common.exceptions import NoSuchElementException 
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import JavascriptException
import pyperclip
import openpyxl
from openpyxl.styles import PatternFill
import traceback
import os
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill
# ----------------------------------------------------------------------------------------------

navegador = None
planTemp = None
sheetPlanTemp = None

def iniciar_navegador(com_debugging_remoto=True):
    #chrome_driver_path = ChromeDriverManager().install()
    chrome_driver_path = r"C:\Users\mateus.silva\.wdm\drivers\chromedriver\win64\131.0.6778.204\chromedriver-win32\chromedriver.exe"
    chrome_driver_executable = os.path.join(os.path.dirname(chrome_driver_path), 'chromedriver.exe')
    
    #print(f"ChromeDriver path: {chrome_driver_executable}")
    if not os.path.isfile(chrome_driver_executable):
        raise FileNotFoundError(f"O ChromeDriver não foi encontrado em {chrome_driver_executable}")

    service = Service(executable_path=chrome_driver_executable)
    
    chrome_options = Options()
    if com_debugging_remoto:
        remote_debugging_port = 9222
        chrome_options.add_experimental_option("debuggerAddress", f"localhost:{remote_debugging_port}")
    
    navegador = webdriver.Chrome(service=service, options=chrome_options)
    return navegador

navegador = iniciar_navegador()

def temp():
    file_path = "Robo_PGFN_excel\\temp.xlsx"

    if os.path.exists(file_path):
        os.remove(file_path)


    planTemp = openpyxl.Workbook()
    sheetPlanTemp = planTemp.active

    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    planTemp.save(file_path)
    print(f"Novo arquivo temp criado com sucesso!")

    return planTemp, sheetPlanTemp

planTemp, sheetPlanTemp = temp()


# PGFN

def verificar_url_PGFN(navegador):
    abas = navegador.window_handles

    for aba in abas:
        navegador.switch_to.window(aba)
        url = navegador.current_url

        if 'regularize.pgfn.gov.br/consultaDividas' in url:
            print("URL PGFN encontrada")
            #navegador.refresh()
            WebDriverWait(navegador, 10).until(lambda navegador: navegador.execute_script('return document.readyState') == 'complete')
            return True
        elif 'regularize.pgfn.gov.br' in url or 'regularize.pgfn.gov.br/consultaDividas/sida/detalhar' in url:
            print("URL PGFN, quase correta encontrada, ajustando URL")
            navegador.get('https://www.regularize.pgfn.gov.br/consultaDividas')
            return True  

    print("Nenhuma aba com a URL correta foi encontrada")
    return False

def percorrer_inscricoes(navegador):
    verificar_url_PGFN(navegador)
    a, g, i = 1, 1, 1

    while True:
        filtro(navegador, a, g)
    
        try:
            WebDriverWait(navegador, 90).until(lambda navegador: navegador.execute_script('return document.readyState') == 'complete')

            try:
                aba = WebDriverWait(navegador, 10).until(
                    EC.presence_of_element_located(
                        (By.XPATH, f'/html/body/app-root/div/div[2]/app-consulta-divida/main/div/div/div/tabset/ul/li[{a}]/a/span[2]')))
            except TimeoutException:
                print("Não há mais abas")
                break

            try:
                grupo = WebDriverWait(navegador, 10).until(
                    EC.presence_of_element_located(
                        (By.XPATH, f'/html/body/app-root/div/div[2]/app-consulta-divida/main/div/div/div/tabset/div/tab[{a}]/app-consulta-divida-aba/app-consulta-divida-grupo[{g}]/div[1]/div/div[1]')))
            except TimeoutException:
                print("Não há mais grupos")
                a += 1
                g = 1
                i = 1
                continue

            if g == 4:
                g = 1
                i = 1
                a += 1
                navegador.execute_script("arguments[0].click();", aba)
                continue

            navegador.execute_script("arguments[0].click();", aba)
            abaIncricao = aba.text
            if '(0)' in abaIncricao:
                g = 1
                i = 1
                a += 1
                navegador.execute_script("arguments[0].click();", aba)
                continue

            tipoIncricao = grupo.text
            if 'Extinta' in tipoIncricao:
                g = 1
                i = 1
                a += 1
                navegador.execute_script("arguments[0].click();", aba)
                continue



            filtro(navegador, a, g)

            inscricoes = navegador.find_elements(By.XPATH, f'/html/body/app-root/div/div[2]/app-consulta-divida/main/div/div/div/tabset/div/tab[{a}]/app-consulta-divida-aba/app-consulta-divida-grupo[{g}]/div[2]/div[3]/table/tbody/tr[{i}]/td[2]/a')
            
            if not inscricoes:
                g += 1
                i = 1
                continue


            inscricao = inscricoes[0]
            numInscricao = inscricao.text

            print("===========================================================")
            print(f'{a}º | Aba       -> {abaIncricao}')
            print(f'{g}º | Grupo     -> {tipoIncricao}')
            print(f'{i}º | Inscricao -> {numInscricao}')

            extrair_informacao(inscricao, abaIncricao)
            i += 1
        except TimeoutException:
            traceback.print_exc()
            return
        except Exception:
            print("erro ao percorrer a lista!")
            traceback.print_exc()
            break

def filtro(navegador, a, g):
    try:
        WebDriverWait(navegador, 90).until(lambda navegador: navegador.execute_script('return document.readyState') == 'complete')

        filtroQntPaginas = WebDriverWait(navegador, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, f'/html/body/app-root/div/div[2]/app-consulta-divida/main/div/div/div/tabset/div/tab[{a}]/app-consulta-divida-aba/app-consulta-divida-grupo[{g}]/div[2]/div[3]/div[1]/div[2]/select')))

        filtrarTodasPaginas = navegador.find_element(By.XPATH, f'/html/body/app-root/div/div[2]/app-consulta-divida/main/div/div/div/tabset/div/tab[{a}]/app-consulta-divida-aba/app-consulta-divida-grupo[{g}]/div[2]/div[3]/div[1]/div[2]/select/option[6]')

        Ofiltro = Select(filtroQntPaginas)
        filtroSelecionado = Ofiltro.first_selected_option
        filtroAplicado = filtroSelecionado.text

        if filtroAplicado != 'Todas':
            filtrarTodasPaginas.click()

        else:
            pass
    except (TimeoutException, JavascriptException):
        pass
    except Exception as e:
        print("erro no filtro!")
        traceback.print_exc()

def extrair_informacao(inscricao, abaIncricao):
    tentativa = 1
    while tentativa < 3:
        try:
            WebDriverWait(navegador, 60).until(lambda navegador: navegador.execute_script('return document.readyState') == 'complete')
            navegador.execute_script("arguments[0].click();", inscricao)
            
            if not 'FGTS' in abaIncricao:
                try:
                    abaRelatorioDetalhado = WebDriverWait(navegador, 10).until(
                        EC.presence_of_element_located((By.XPATH, '/html/body/app-root/div/div[2]/app-inscricao/main/tabset/ul/li[2]/a')))
                except TimeoutException:
                    # Tentar encontrar o elemento alternativo
                    abaRelatorioDetalhado = WebDriverWait(navegador, 10).until(
                        EC.presence_of_element_located((By.XPATH, '/html/body/app-root/div/div[2]/app-debcad/main/tabset/ul/li[2]/a')))
                navegador.execute_script("arguments[0].click();", abaRelatorioDetalhado)

                campoFiltrarTodos = WebDriverWait(navegador, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="todosCheck"]')))
                navegador.execute_script("arguments[0].click();", campoFiltrarTodos)
                

                try:
                    botaoGerarRelatorio = WebDriverWait(navegador, 10).until(
                        EC.presence_of_element_located((By.XPATH, '/html/body/app-root/div/div[2]/app-inscricao/main/tabset/div/tab[2]/fieldset/div/button[1]')))
                except TimeoutException:
                    # Tentar encontrar o elemento alternativo
                    botaoGerarRelatorio = WebDriverWait(navegador, 10).until(
                        EC.presence_of_element_located((By.XPATH, '/html/body/app-root/div/div[2]/app-debcad/main/tabset/div/tab[2]/fieldset/div/button[1]')))
                navegador.execute_script("arguments[0].click();", botaoGerarRelatorio)

            pyperclip.copy('')
            time.sleep(0.1)

            actions = ActionChains(navegador)
            time.sleep(2.2)
            actions.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            time.sleep(1)
            actions.key_down(Keys.CONTROL).send_keys('c').key_up(Keys.CONTROL).perform()
            time.sleep(1)

            informacoes = pyperclip.paste()

            adicionar_info(informacoes)

            tirarSelecao = navegador.find_element(By.XPATH, '/html/body/app-root/div').click()
            #navegador.execute_script("arguments[0].click();", tirarSelecao)

            xpaths = [
                '/html/body/app-root/div/div[2]/app-inscricao/main/div[3]/button[1]',
                '/html/body/app-root/div/div[2]/app-debcad/main/div[4]/button[1]',
                '/html/body/app-root/div/div[2]/app-fgts/main/div[5]/button[1]'
            ]
            botaoVoltar = None

            for xpath in xpaths:
                try:
                    botaoVoltar = WebDriverWait(navegador, 10).until(
                        EC.presence_of_element_located((By.XPATH, xpath))
                    )
                    break
                except TimeoutException:
                    continue

            
            navegador.execute_script("arguments[0].click();", botaoVoltar)


            #navegador.back()
            WebDriverWait(navegador, 10).until(lambda navegador: navegador.execute_script('return document.readyState') == 'complete')
            break
        except TimeoutException:
            tentativa += 1
            print(f"tentativas: {tentativa}")
            return
        except Exception:
            traceback.print_exc()
            break

def adicionar_info(informacoes):

    # Dividir as informações em linhas
    linhas = informacoes.split('\n')

    # Abrir o arquivo Excel

    planTemp = openpyxl.load_workbook("Robo_PGFN_excel\\temp.xlsx")
    sheetPlanTemp = planTemp.active

    # Obter a próxima linha vazia na planilha
    proxima_linha = sheetPlanTemp.max_row + 1

    # Adicionar os dados na planilha, mantendo a formatação original
    for linha in linhas:
        # Dividir os dados da linha, se necessário
        dados = linha.split("\t")  # Use o delimitador correto, como "\t" para tabulação
        for j, valor in enumerate(dados, start=1):
            sheetPlanTemp.cell(row=proxima_linha, column=j).value = valor
        proxima_linha += 1


    #formulas(planTemp)

    # Salvar as alterações na planilha
    planTemp.save('Robo_PGFN_excel\\temp.xlsx')
    #criar_plan_final(planTemp)

def formulas(planTemp):
    planTemp = openpyxl.load_workbook("Robo_PGFN_excel\\temp.xlsx")

    paginaPlan = planTemp['Sheet']
    ultima_linha = paginaPlan.max_row
    linha = 2

    fill_azul_claro = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  
    fill_roxo_claro = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")  
    fill_vermelho_claro = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")  
    fill_verde_claro = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  
    fill_laranja_claro = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")



    # PADRAO
    paginaPlan['AT1'] = "SITUACAO"
    paginaPlan['AT1'].fill = fill_laranja_claro
    formula = f'=IF(ISNUMBER(SEARCH("Situação da inscrição:",A{linha})),MID(A{linha},24,50),AT{linha-1})&IF(0,"","")'
    paginaPlan[f'AT{linha}'] = formula 
    paginaPlan[f'AT{ultima_linha}'] = 0 

    paginaPlan['I1'] = "Nº_INSCRICAO"
    paginaPlan['I1'].fill = fill_laranja_claro
    formula = f'=IF(ISNUMBER(SEARCH("Nº inscrição",A{linha})),MID(A{linha},15,18),I{linha-1})&IF(0,"","")'
    paginaPlan[f'I{linha}'] = formula 
    paginaPlan[f'I{ultima_linha}'] = 0 

    paginaPlan['J1'] = "CNPJ"
    paginaPlan['J1'].fill = fill_laranja_claro
    formula = f'=IF(ISNUMBER(SEARCH("CPF/CNPJ:",A{linha})),MID(A{linha},10,19),J{linha-1})'
    paginaPlan[f'J{linha}'] = formula
    paginaPlan[f'J{ultima_linha}'] = 0

    paginaPlan['K1'] = "NOME"
    paginaPlan['K1'].fill = fill_laranja_claro
    formula = f'=IF(ISNUMBER(SEARCH("Devedor principal",A{linha})),MID(A{linha},19,50),K{linha-1})'
    paginaPlan[f'K{linha}'] = formula
    paginaPlan[f'K{ultima_linha}'] = 0

    paginaPlan['L1'] = "DT_INSCRICAO"
    paginaPlan['L1'].fill = fill_laranja_claro
    formula = f'=IF(ISNUMBER(SEARCH("Data da inscrição",A{linha})),MID(A{linha},20,10),L{linha-1})'
    paginaPlan[f'L{linha}'] = formula
    paginaPlan[f'L{ultima_linha}'] = 0

    paginaPlan['M1'] = "RECEITA_DA_DIVIDA"
    paginaPlan['M1'].fill = fill_laranja_claro
    formula = f'=IF(ISNUMBER(SEARCH("Receita da dívida",A{linha})),MID(A{linha},20,50),M{linha-1})'
    paginaPlan[f'M{linha}'] = formula
    paginaPlan[f'M{ultima_linha}'] = 0

    paginaPlan['N1'] = "Nº_PROCESSO"
    paginaPlan['N1'].fill = fill_laranja_claro
    formula = f'=IF(ISNUMBER(SEARCH("Nº do processo administrativo",A{linha})),MID(A{linha},32,50),N{linha-1})'
    paginaPlan[f'N{linha}'] = formula
    paginaPlan[f'N{ultima_linha}'] = 0

    paginaPlan['O1'] = "VL_PRINCIPAL"
    paginaPlan['O1'].fill = fill_verde_claro
    formula = f'=IF(ISNUMBER(SEARCH("Valor principal",A{linha+2})),A{linha+3},IF(ISNUMBER(SEARCH("R$",V{linha+2})),O{linha-1},""))'
    paginaPlan[f'O{linha+4}'] = formula
    paginaPlan[f'O{ultima_linha}'] = 0

    paginaPlan['P1'] = "VL_MULTA"
    paginaPlan['P1'].fill = fill_verde_claro
    formula = f'=IF(ISNUMBER(SEARCH("Multa",B{linha+2})),B{linha+3},IF(ISNUMBER(SEARCH("R$",V{linha+2})),P{linha-1},""))'
    paginaPlan[f'P{linha+4}'] = formula
    paginaPlan[f'P{ultima_linha}'] = 0

    paginaPlan['Q1'] = "VL_JUROS"
    paginaPlan['Q1'].fill = fill_verde_claro
    formula = f'=IF(ISNUMBER(SEARCH("Juros",C{linha+2})),C{linha+3},IF(ISNUMBER(SEARCH("R$",V{linha+2})),Q{linha-1},""))'
    paginaPlan[f'Q{linha+4}'] = formula
    paginaPlan[f'Q{ultima_linha}'] = 0

    paginaPlan['R1'] = "VL_ENCARGO"
    paginaPlan['R1'].fill = fill_verde_claro
    formula = f'=IF(ISNUMBER(SEARCH("Encargo legal",D{linha+2})),D{linha+3},IF(ISNUMBER(SEARCH("R$",V{linha+2})),R{linha-1},""))'
    paginaPlan[f'R{linha+4}'] = formula
    paginaPlan[f'R{ultima_linha}'] = 0

    paginaPlan['S1'] = "VL_TOTAL_CONSOLIDADO"
    paginaPlan['S1'].fill = fill_verde_claro
    formula = f'=IF(ISNUMBER(SEARCH("Valor total consolidado",E{linha+2})),E{linha+3},IF(ISNUMBER(SEARCH("R$",V{linha+2})),S{linha-1},""))'
    paginaPlan[f'S{linha+4}'] = formula
    paginaPlan[f'S{ultima_linha}'] = 0

    paginaPlan['T1'] = "DT_INICIO_APURACAO"
    paginaPlan['T1'].fill = fill_verde_claro
    formula = f'=IF(ISNUMBER(SEARCH("Exibir o detalhamento",G{linha+4})),A{linha+3},"")'
    paginaPlan[f'T{linha}'] = formula
    paginaPlan[f'T{ultima_linha}'] = 0

    paginaPlan['U1'] = "DT_FIM_APURACAO"
    paginaPlan['U1'].fill = fill_verde_claro
    formula = f'=IF(ISNUMBER(SEARCH("Exibir o detalhamento",G{linha+4})),A{linha+4},"")'
    paginaPlan[f'U{linha}'] = formula
    paginaPlan[f'U{ultima_linha}'] = 0

    paginaPlan['V1'] = "VL_DEB"
    paginaPlan['V1'].fill = fill_verde_claro
    formula = f'=IF(ISNUMBER(SEARCH("Exibir o detalhamento",G{linha+4})),B{linha+4},"")'
    paginaPlan[f'V{linha}'] = formula
    paginaPlan[f'V{ultima_linha}'] = 0

    paginaPlan['W1'] = "VL_REMANESCENTE"
    paginaPlan['W1'].fill = fill_verde_claro
    formula = f'=IF(ISNUMBER(SEARCH("Exibir o detalhamento",G{linha+4})),C{linha+4},"")'
    paginaPlan[f'W{linha}'] = formula
    paginaPlan[f'W{ultima_linha}'] = 0

    paginaPlan['X1'] = "VL_MULTA_MORA"
    paginaPlan['X1'].fill = fill_verde_claro
    formula = f'=IF(ISNUMBER(SEARCH("Exibir o detalhamento",G{linha+4})),D{linha+4},"")'
    paginaPlan[f'X{linha}'] = formula
    paginaPlan[f'X{ultima_linha}'] = 0

    paginaPlan['Y1'] = "NAT_DEB !"
    paginaPlan['Y1'].fill = fill_verde_claro
    formula = f'=IF(ISNUMBER(SEARCH("Exibir o detalhamento",G{linha+4})),E{linha+4},"")'
    paginaPlan[f'Y{linha}'] = formula
    paginaPlan[f'Y{ultima_linha}'] = 0

    # FGTS
    paginaPlan['Z1'] = "FGTS DT_AJUIZAMENTO"
    paginaPlan['Z1'].fill = fill_azul_claro
    formula = f'=IF(ISNUMBER(SEARCH("Data do ajuizamento",A{linha})),MID(A{linha},22,30),"")'
    paginaPlan[f'Z{linha}'] = formula
    paginaPlan[f'Z{ultima_linha}'] = 0

    paginaPlan['AA1'] = "FGTS Nº PROCESSO"
    paginaPlan['AA1'].fill = fill_azul_claro
    formula = f'=IF(ISNUMBER(SEARCH("Nº único do processo",A{linha+1})),MID(A{linha+1},32,30),"")'
    paginaPlan[f'AA{linha}'] = formula
    paginaPlan[f'AA{ultima_linha}'] = 0

    paginaPlan['AB1'] = "FGTS VL_TOTAL_CONSOLIDADO"
    paginaPlan['AB1'].fill = fill_azul_claro
    formula = f'=IF(ISNUMBER(SEARCH("Valor total consolidado",A{linha+4})),A{linha+5},"")'
    paginaPlan[f'AB{linha}'] = formula
    paginaPlan[f'AB{ultima_linha}'] = 0


    # EM COBRANCA
    paginaPlan['AC1'] = "EM_COBRANÇA VL_PRINCIPAL"
    paginaPlan['AC1'].fill = fill_vermelho_claro
    # =O2&IF(0,"","")
    formula = f'=O2&IF(0,"","")'
    paginaPlan[f'AC{linha}'] = formula
    paginaPlan[f'AC{ultima_linha}'] = 0

    paginaPlan['AD1'] = "EM_COBRANÇA VL_MULTA_MORA"
    paginaPlan['AD1'].fill = fill_vermelho_claro
    # =P2&IF(0,"","")
    formula = f'=P2&IF(0,"","")'
    paginaPlan[f'AD{linha}'] = formula
    paginaPlan[f'AD{ultima_linha}'] = 0
    
    paginaPlan['AE1'] = "EM_COBRANÇA VL_MULTA_DE_OFICIO"
    paginaPlan['AE1'].fill = fill_vermelho_claro
    # =SE(ÉNÚM(LOCALIZAR("Multa";C1));C2;"")
    formula = f'=IF(ISNUMBER(SEARCH("Multa",C{linha-1})),C{linha},"")'
    paginaPlan[f'AE{linha+1}'] = formula
    paginaPlan[f'AE{ultima_linha}'] = 0

    paginaPlan['AF1'] = "EM_COBRANÇA VL_MULTA_ISOLADA"
    paginaPlan['AF1'].fill = fill_vermelho_claro
    # =SE(ÉNÚM(LOCALIZAR("Multa";D1));D2;""))
    formula = f'=IF(ISNUMBER(SEARCH("Multa",D{linha-1})),D{linha},"")'
    paginaPlan[f'AF{linha+1}'] = formula
    paginaPlan[f'AF{ultima_linha}'] = 0

    paginaPlan['AG1'] = "EM_COBRANÇA VL_JUROS_MORA"
    paginaPlan['AG1'].fill = fill_vermelho_claro
    # =SE(ÉNÚM(LOCALIZAR("Juros de mora";E1));E2;"")
    formula = f'=IF(ISNUMBER(SEARCH("Juros de mora",E{linha-1})),E{linha},"")'
    paginaPlan[f'AG{linha+1}'] = formula
    paginaPlan[f'AG{ultima_linha}'] = 0

    paginaPlan['AH1'] = "EM_COBRANÇA VL_ENCARGO_LEGAL"
    paginaPlan['AH1'].fill = fill_vermelho_claro
    # =SE(ÉNÚM(LOCALIZAR("Encargo legal";F1));F2;"")
    formula = f'=IF(ISNUMBER(SEARCH("Encargo legal",F{linha-1})),F{linha},"")'
    paginaPlan[f'AH{linha+1}'] = formula
    paginaPlan[f'AH{ultima_linha}'] = 0

    paginaPlan['AI1'] = "EM_COBRANÇA VL_HONORARIOS"
    paginaPlan['AI1'].fill = fill_vermelho_claro
    # =SE(ÉNÚM(LOCALIZAR("Honorários";G1));G2;"")
    formula = f'=IF(ISNUMBER(SEARCH("Honorários",G{linha-1})),G{linha},"")'
    paginaPlan[f'AI{linha+1}'] = formula
    paginaPlan[f'AI{ultima_linha}'] = 0

    paginaPlan['AJ1'] = "EM_COBRANÇA VL_TOTAL"
    paginaPlan['AJ1'].fill = fill_vermelho_claro
    formula = f'=IF(ISNUMBER(SEARCH("Valor total",H{linha-1})),H{linha},"")' 
    paginaPlan[f'AJ{linha+1}'] = formula
    paginaPlan[f'AJ{ultima_linha}'] = 0

    paginaPlan['AK1'] = "EM_COBRANÇA COMPETENCIA"
    paginaPlan['AK1'].fill = fill_vermelho_claro
    formula = f'=IF(ISNUMBER(SEARCH("/",B{linha+2})),B{linha+2},"")'
    paginaPlan[f'AK{linha}'] = formula
    paginaPlan[f'AK{ultima_linha}'] = 0

    paginaPlan['AL1'] = "EM_COBRANÇA VL_ITEM"
    paginaPlan['AL1'].fill = fill_vermelho_claro
    formula = f'=IF(ISNUMBER(SEARCH("/",B{linha+2})),C{linha+2},"")'
    paginaPlan[f'AL{linha}'] = formula
    paginaPlan[f'AL{ultima_linha}'] = 0

    paginaPlan['AM1'] = "EM_COBRANÇA SALDO_DEVEDOR"
    paginaPlan['AM1'].fill = fill_vermelho_claro
    formula = f'=IF(ISNUMBER(SEARCH("/",B{linha+2})),D{linha+2},"")'
    paginaPlan[f'AM{linha}'] = formula
    paginaPlan[f'AM{ultima_linha}'] = 0


    # PAGAMENTOS
    paginaPlan['AN1'] = "PAGAMENTOS DT_INICIO_ARRECADACAO"
    paginaPlan['AN1'].fill = fill_roxo_claro
    # =SE(E(ÉNÚM(LOCALIZAR("/";A9));ÉNÚM(LOCALIZAR("R$";B9));G9="");A8;"")
    formula = f'=IF(AND(ISNUMBER(SEARCH("/",A{linha+7})),ISNUMBER(SEARCH("R$",B{linha+7})),G{linha+7}=""),A{linha+6},"")'
    paginaPlan[f'AN{linha}'] = formula
    paginaPlan[f'AN{ultima_linha}'] = 0

    paginaPlan['AO1'] = "PAGAMENTOS DT_FIM_ARRECADACAO"
    paginaPlan['AO1'].fill = fill_roxo_claro
    # =SE(E(ÉNÚM(LOCALIZAR("/";A9));ÉNÚM(LOCALIZAR("R$";B9));G9="");A9;"")
    formula = f'=IF(AND(ISNUMBER(SEARCH("/",A{linha+7})),ISNUMBER(SEARCH("R$",B{linha+7})),G{linha+7}=""),A{linha+7},"")'
    paginaPlan[f'AO{linha}'] = formula
    paginaPlan[f'AO{ultima_linha}'] = 0

    paginaPlan['AP1'] = "PAGAMENTOS VL_APROPRIADO"
    paginaPlan['AP1'].fill = fill_roxo_claro
    # =SE(E(ÉNÚM(LOCALIZAR("/";A9));ÉNÚM(LOCALIZAR("R$";B9));G9="");B9;"")
    formula = f'=IF(AND(ISNUMBER(SEARCH("/",A{linha+7})),ISNUMBER(SEARCH("R$",B{linha+7})),G{linha+7}=""),B{linha+7},"")'
    paginaPlan[f'AP{linha}'] = formula
    paginaPlan[f'AP{ultima_linha}'] = 0

    paginaPlan['AQ1'] = "PAGAMENTOS REFERENCIA"
    paginaPlan['AQ1'].fill = fill_roxo_claro
    # =SE(E(ÉNÚM(LOCALIZAR("/";A9));ÉNÚM(LOCALIZAR("R$";B9));G9="");C9;"")
    formula = f'=IF(AND(ISNUMBER(SEARCH("/",A{linha+7})),ISNUMBER(SEARCH("R$",B{linha+7})),G{linha+7}=""),C{linha+7},"")'
    paginaPlan[f'AQ{linha}'] = formula
    paginaPlan[f'AQ{ultima_linha}'] = 0

    paginaPlan['AR1'] = "PAGAMENTOS Nº GUIA !"
    paginaPlan['AR1'].fill = fill_roxo_claro
    # =SE(E(ÉNÚM(LOCALIZAR("/";A9));ÉNÚM(LOCALIZAR("R$";B9));G9="");D9;"")
    formula = f'=IF(AND(ISNUMBER(SEARCH("/",A{linha+7})),ISNUMBER(SEARCH("R$",B{linha+7})),G{linha+7}=""),D{linha+7},"")'
    paginaPlan[f'AR{linha}'] = formula
    paginaPlan[f'AR{ultima_linha}'] = 0

    paginaPlan['AS1'] = "PAGAMENTOS TIPO_CREDITO"
    paginaPlan['AS1'].fill = fill_roxo_claro
    #=SE(E(ÉNÚM(LOCALIZAR("/";A9));ÉNÚM(LOCALIZAR("R$";B9));G9="");E9;"")
    formula = f'=IF(AND(ISNUMBER(SEARCH("/",A{linha+7})),ISNUMBER(SEARCH("R$",B{linha+7})),G{linha+7}=""),E{linha+7},"")'
    paginaPlan[f'AS{linha}'] = formula
    paginaPlan[f'AS{ultima_linha}'] = 0

    planTemp.save('Robo_PGFN_excel\\temp.xlsx')

def renomear_plan():
    verificar_url_PGFN(navegador)
    WebDriverWait(navegador, 90).until(lambda navegador: navegador.execute_script('return document.readyState') == 'complete')

    abaPerfil = navegador.find_element(By.XPATH, '//*[@id="navbar"]/div/div[1]/div/button')
    navegador.execute_script("arguments[0].click();", abaPerfil)

    campoNome = navegador.find_element(By.XPATH, '//*[@id="userInfoMenu"]/div/div/b')
    nomeEmpresa = campoNome.text

    caminho_pasta = Path("C:\\VS_CODE\\Robo_PGFN_excel")
    arquivo = "temp.xlsx"
    caminho_arquivo = caminho_pasta / arquivo
    try:
        if caminho_arquivo.is_file():
            extensao = caminho_arquivo.suffix
            novo_caminho = caminho_pasta / (f"Relatório PGFN - {nomeEmpresa}{extensao}")
            caminho_arquivo.rename(novo_caminho)

            print(f"Arquivo {arquivo} renomeado para Relatório PGFN - {nomeEmpresa}{extensao}")
        else:
            print("Arquivo não encontrado")
    except Exception as e:
        print(f"Erro ao renomear arquivo:\n{e}")

# ----------------------------------------------------------------------------------------------
print("iniciando...")
time.sleep(2)
tempo_inicio_total = datetime.now()
percorrer_inscricoes(navegador)
print("aplicando formulas")
formulas(planTemp)
renomear_plan()

tempo_fim_total = datetime.now()
tempo_execução_total = tempo_fim_total - tempo_inicio_total
tempo_formatado = str(tempo_execução_total)
largura_total = 28
largura_interna = largura_total - 11 

linha_superior = "|¯¯ TEMPO DE EXECUÇÃO ¯¯|".center(largura_total)
linha_inferior = f"|__ {tempo_formatado.center(largura_interna)} __|".center(largura_total)
print(linha_superior)
print(linha_inferior)