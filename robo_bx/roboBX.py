import cv2
import numpy as np
import difflib
from pywinauto import Desktop
import pyautogui
import time
import random
import mousekey
import pytesseract
import pygetwindow as gw
from PIL import Image
from PIL import ImageGrab
from datetime import datetime
import shutil
import pyscreeze
import os
import subprocess
from pathlib import Path
import json
import re
import openpyxl
import openpyxl.workbook
from tqdm import tqdm
import time
import psutil
import ctypes
import win32con
import win32api

#----------------------------------------------FUNÇÕES----------------------------------------------------
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
planilha_caminhos = openpyxl.load_workbook(r"robo_bx\BasesNovas.xlsx") #ARQUIVO XLSX 
#planilha_caminhos = openpyxl.load_workbook(r"C:\Users\gabriel.alvise\Desktop\EMPRESAS\MATEUS\baixarArq.xlsx")
pagiCaminhos = planilha_caminhos['Planilha1'] 

def buscar_e_clicar(texto_busca, ocorrencia=1, horizontal=10, vertical=10, click=1, coordenadas=None, max_tentativas=50):

    def preprocessar_imagem(imagem):
        imagem_cinza = cv2.cvtColor(np.array(imagem), cv2.COLOR_BGR2GRAY)
        _, imagem_binaria = cv2.threshold(imagem_cinza, 150, 255, cv2.THRESH_BINARY)
        return Image.fromarray(imagem_binaria)
    
    def tentar_deteccao(imagem, config):
        try:
            return pytesseract.image_to_data(imagem, lang='por', output_type=pytesseract.Output.DICT, config=config)
        except Exception as e:
            print(f"Erro na detecção: {e}")
            return None
        
    def normalizar_texto(texto):
        return ''.join(c for c in texto if c.isalnum()).lower()
    
    def palavras_similares(palavra_detectada, palavra_busca):
        similaridade = difflib.SequenceMatcher(None, normalizar_texto(palavra_detectada), normalizar_texto(palavra_busca)).ratio()
        return similaridade > 0.8 

    tentativa = 0
    encontrado = False

    if coordenadas:
        x, y = coordenadas
        bbox = (x - 50, y - 50, x + 50, y + 50)
    else:
        bbox = None 

    while tentativa < max_tentativas and not encontrado:
        print(f"Tentativa {tentativa + 1} de {max_tentativas} para encontrar '{texto_busca}'")

        screenshot = ImageGrab.grab(bbox=bbox)  
        if coordenadas:
            print(f'Capturando área: {coordenadas}')
        else:
            print('Capturando tela inteira.')

        largura_imagem, altura_imagem = screenshot.size  

        imagem_processada = preprocessar_imagem(screenshot)
        config = '--oem 3 --psm 6'
        texto_tela = tentar_deteccao(imagem_processada, config)

        if not texto_tela:
            print(f'Não foi possível detectar o texto.')
            tentativa += 1
            continue
        
        palavras_detectadas = texto_tela['text']
        coordenadas_detectadas = list(zip(texto_tela['left'], texto_tela['top'], texto_tela['width'], texto_tela['height']))

        # Verifica as coordenadas de cada palavra
        """
        print("Palavras e suas coordenadas detectadas:")
        for i, palavra in enumerate(palavras_detectadas):
            if palavra.strip() != "":
                print(f"Palavra: '{palavra}', Coordenadas: ({texto_tela['left'][i]}, {texto_tela['top'][i]})")
        """

        ocorrencias_encontradas = [
            (x, y, largura, altura)
            for i, palavra in enumerate(palavras_detectadas)
            if palavras_similares(palavra, texto_busca.split()[0])
            for x, y, largura, altura in [coordenadas_detectadas[i]]
        ]

        if len(ocorrencias_encontradas) >= ocorrencia:
            x, y, largura, altura = ocorrencias_encontradas[ocorrencia - 1]
            encontrado = True
        else:
            print(f'A ocorrência {ocorrencia} da palavra "{texto_busca}" não foi encontrada.')
            tentativa += 1

    if encontrado:
        if 0 <= x <= largura_imagem and 0 <= y <= altura_imagem:
            pyautogui.moveTo(x + horizontal, y + vertical)
            if click == 2:
                pyautogui.doubleClick()
            else:
                pyautogui.click()
            print(f'Clique na ocorrência {ocorrencia} da palavra "{texto_busca}" nas coordenadas ({x + horizontal}, {y + vertical})!')
        else:
            print(f'Coordenadas ({x}, {y}) estão fora dos limites da tela.')
    else:
        print(f'Não foi possível encontrar a palavra "{texto_busca}" após {max_tentativas} tentativas.')
    
    return encontrado

def buscar_e_clicar_todas(texto_busca, horizontal=10, vertical=10, click=1, coordenadas=None, max_tentativas=50):
    import cv2
    import numpy as np
    import difflib
    import pyautogui
    from PIL import Image, ImageGrab
    from pytesseract import pytesseract

    def preprocessar_imagem(imagem):
        imagem_cinza = cv2.cvtColor(np.array(imagem), cv2.COLOR_BGR2GRAY)
        _, imagem_binaria = cv2.threshold(imagem_cinza, 150, 255, cv2.THRESH_BINARY)
        return Image.fromarray(imagem_binaria)
    
    def tentar_deteccao(imagem, config):
        try:
            return pytesseract.image_to_data(imagem, lang='por', output_type=pytesseract.Output.DICT, config=config)
        except Exception as e:
            print(f"Erro na detecção: {e}")
            return None
        
    def normalizar_texto(texto):
        return ''.join(c for c in texto if c.isalnum()).lower()
    
    def palavras_similares(palavra_detectada, palavra_busca):
        similaridade = difflib.SequenceMatcher(None, normalizar_texto(palavra_detectada), normalizar_texto(palavra_busca)).ratio()
        return similaridade > 0.8 

    tentativa = 0
    encontrado = False
    ocorrencias_encontradas = []

    if coordenadas:
        x, y = coordenadas
        bbox = (x - 50, y - 50, x + 50, y + 50)
    else:
        bbox = None 

    while tentativa < max_tentativas:
        print(f"Tentativa {tentativa + 1} de {max_tentativas} para encontrar todas as ocorrências de '{texto_busca}'")

        screenshot = ImageGrab.grab(bbox=bbox)  
        if coordenadas:
            print(f'Capturando área: {coordenadas}')
        else:
            print('Capturando tela inteira.')

        largura_imagem, altura_imagem = screenshot.size  

        imagem_processada = preprocessar_imagem(screenshot)
        config = '--oem 3 --psm 6'
        texto_tela = tentar_deteccao(imagem_processada, config)

        if not texto_tela:
            print(f'Não foi possível detectar o texto.')
            tentativa += 1
            continue
        
        palavras_detectadas = texto_tela['text']
        coordenadas_detectadas = list(zip(texto_tela['left'], texto_tela['top'], texto_tela['width'], texto_tela['height']))

        # Adiciona todas as ocorrências encontradas
        novas_ocorrencias = [
            (x, y, largura, altura)
            for i, palavra in enumerate(palavras_detectadas)
            if palavras_similares(palavra, texto_busca.split()[0])
            for x, y, largura, altura in [coordenadas_detectadas[i]]
        ]
        
        ocorrencias_encontradas.extend(novas_ocorrencias)
        
        if novas_ocorrencias:
            print(f'{len(novas_ocorrencias)} ocorrências encontradas nesta tentativa.')
            encontrado = True
            break  # Pode parar ao encontrar todas as ocorrências
        else:
            print(f'Nenhuma ocorrência encontrada nesta tentativa.')
            tentativa += 1

    if encontrado:
        for x, y, largura, altura in ocorrencias_encontradas:
            if 0 <= x <= largura_imagem and 0 <= y <= altura_imagem:
                pyautogui.moveTo(x + horizontal, y + vertical)
                if click == 2:
                    pyautogui.doubleClick()
                else:
                    pyautogui.click()
                print(f'Clicado em "{texto_busca}" nas coordenadas ({x + horizontal}, {y + vertical})!')
            else:
                print(f'Coordenadas ({x}, {y}) estão fora dos limites da tela.')
    else:
        print(f'Não foi possível encontrar a palavra "{texto_busca}" após {max_tentativas} tentativas.')
    
    return encontrado

def procurar_imagem(nome_arquivo, confidence=0.8, region=None, max_tentativas=60, horizontal=0, vertical=0, dx=0, dy=0, acao='clicar', clicks=1, ocorrencia=1, delay_tentativa=1):
    mkey = mousekey.MouseKey()

    def click(x, y):
        pyautogui.click(x, y)

    def doubleClick(x, y):
        pyautogui.doubleClick(x, y)

    def coordenada(x, y):
        #print(f'Coordenadas da imagem: ({x}, {y})')
        print()

    def moveMouse(x, y, variationx=(-5, 5), variationy=(-5, 5), up_down=(0.2, 0.3), min_variation=-10, max_variation=10, use_every=4, sleeptime=(0.009, 0.019), linear=90):
        mkey.left_click_xy_natural(
            int(x) - random.randint(*variationx),
            int(y) - random.randint(*variationy),
            delay=random.uniform(*up_down),
            min_variation=min_variation,
            max_variation=max_variation,
            use_every=use_every,
            sleeptime=sleeptime,
            print_coords=True,
            percent=linear,
        )

    def clickDrag(x, y, dx, dy):
        pyautogui.moveTo(x, y)
        pyautogui.mouseDown()
        pyautogui.moveTo(x + dx, y + dy, duration=0.5)
        pyautogui.mouseUp()


    acoesValidas = ['clicar', 'mover clicar', 'clicar arrastar']

    if acao not in acoesValidas:
        raise ValueError(f"Ação inválida: '{acao}'. Escolha entre {acoesValidas}.")

    tentativas = 0
    while tentativas < max_tentativas:
        tentativas += 1
        try:
            imag = list(pyautogui.locateAllOnScreen(nome_arquivo, confidence=confidence, region=region))
            
            if imag:
                if len(imag) >= ocorrencia:
                    img = imag[ocorrencia - 1]  
                    x, y = pyautogui.center(img) 
                    x += horizontal
                    y += vertical

                    match acao:
                        case 'clicar':
                            match clicks:
                                case 0:
                                    coordenada(x, y)
                                case 1:
                                    click(x, y)
                                case 2:
                                    doubleClick(x, y)

                        case 'mover clicar':
                            moveMouse(x, y)
                        
                        case 'clicar arrastar':
                            if dx == 0 and dy == 0:
                                print("Para clicar e arrastar é necessário adicionar os parâmetros (dx, dy)")
                                return False  
                            else:
                                clickDrag(x, y, dx, dy)
                            return True

                    return True
                else:
                    print(f'A ocorrência {ocorrencia} não foi encontrada.')
                    return False

        except pyscreeze.ImageNotFoundException:
            pass
        time.sleep(delay_tentativa)
        
    if acao == 'clicar arrastar' and dx == 0 and dy == 0:
        print(f'Imagem encontrada, porém os parâmetros de arrasto (dx, dy) estão faltando. Não foi possível completar a ação.')
    else:
        print(f'Imagem não encontrada após {max_tentativas} tentativas.')
    return False

def set_high_priority(process_name):
    for proc in psutil.process_iter(attrs=['pid', 'name']):
        if proc.info['name'].lower() == process_name.lower():
            pid = proc.info['pid']
            print(f"Alterando prioridade do processo {process_name} (PID: {pid}) para ALTA...")

            # Abrir o processo com permissões necessárias
            handle = win32api.OpenProcess(win32con.PROCESS_ALL_ACCESS, False, pid)

            # Converter para inteiro para compatibilidade com ctypes
            ctypes.windll.kernel32.SetPriorityClass(int(handle), win32con.HIGH_PRIORITY_CLASS)
            print("Prioridade alterada com sucesso!")
            return True

    print("Processo não encontrado.")
    return False

def login(cnpj):
    titulos = gw.getAllTitles()
    janelasBX = titulos.count('Receitanet BX')

    buscar_e_clicar('MARCIO RODRIGO FRIZZO')
    procurar_imagem(r'C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\btnContribuinte.png')
    procurar_imagem(r'C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\btnProcurador.png')
    time.sleep(0.5)

    procurar_imagem(r'C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\btnCpf.png')
    procurar_imagem(r'C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\btnCnpj.png')
    procurar_imagem(r'C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\inputCnpj.png')

    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.5)
    print(cnpj)
    pyautogui.write(cnpj)
    time.sleep(0.5)
    procurar_imagem(r'C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\btnEntrar.png')

    time.sleep(5)
    
    app = Desktop(backend='uia').window(title='Receitanet BX')
    botaoMaximiza = app.child_window(title='Maximizar', control_type='Button')
    botaoMaximiza.click_input()
    time.sleep(0.5)

def verificarProc(alinha, cnpj):
    while True:
        titulos = gw.getAllTitles()
        contagem = titulos.count('Pesquisando por arquivos...')

        if contagem == 1:
            print(f"Há {contagem} janelas 'Pesquisando arquivos' abertas.")
            time.sleep(1.5)

            titulos = gw.getAllTitles()
            contagemReceitanetBX = titulos.count('Receitanet BX')

            if contagemReceitanetBX > 1:
                print(f"Há {contagemReceitanetBX} janelas 'Receitanet BX' abertas.")

                encontrou_nao_existe_procuracao = False  

                for tentativa in range(3):
                    encontrado = buscar_e_clicar("Não existe procuração", max_tentativas=3)
                    if encontrado:
                        encontrou_nao_existe_procuracao = True
                        pyautogui.press('enter')
                        pagiCaminhos.cell(column=4, row=alinha, value='SEM PROC')
                        planilha_caminhos.save(r'robo_bx\sem_proc.xlsx')
                        print(f'{cnpj} - NÃO EXISTE PROCURAÇÃO')
                        break

                if encontrou_nao_existe_procuracao:
                    break

                encontrou_arquivo = False
                for tentativa in range(3):
                    if procurar_imagem(r"C:\Users\gabriel.alvise\Desktop\ROBOS\BX - Copia\Prints\nenhumarquivo.png", max_tentativas=3):
                        encontrou_arquivo = True
                        pyautogui.press('enter')
                        pagiCaminhos.cell(column=4, row=alinha, value='NENHUM ARQUIVO')
                        planilha_caminhos.save(r'robo_bx\sem_proc.xlsx')
                        print(f'{cnpj} - NENHUM ARQUIVO ENCONTRADO')
                        break

            else:
                pagiCaminhos.cell(column=4, row=alinha, value='TEM PROC')
                planilha_caminhos.save(r'robo_bx\sem_proc.xlsx')
                print(f'{cnpj} - TEM PROCURAÇÃO')


        else:
            print(f"Não há janelas 'Pesquisando arquivos' abertas. Finalizando...")
            break           

def tratarSpedEcf(pastaDestinoEcf, nomeEmpresa):  
    arquivosSeparados = {}
    pastaBackup = r"C:\Users\gabriel.alvise\Documents\BackupECF"
    
    for arquivo in os.listdir(pastaDestinoEcf):
        if arquivo.startswith("SPEDECF") and arquivo.endswith(".txt"):
            partes = arquivo.split('-')
            if len(partes) > 4:
                periodo = partes[2] + '-' + partes[3]
                dataHoraStr = partes[4].replace('.txt', '')

                dataHora = datetime.strptime(dataHoraStr, "%Y%m%d%H%M%S")

                nome_empresa = fr"C:\Users\gabriel.alvise\Documents\BackupECF\{nomeEmpresa}"

                if not os.path.exists(nome_empresa):
                    os.makedirs(nome_empresa)
                    pastaBackup = os.path.join(pastaBackup, nome_empresa)
                else:
                    pastaBackup = os.path.join(pastaBackup, nome_empresa)

                if periodo in arquivosSeparados:
                    if dataHora > arquivosSeparados[periodo]['data_hora']:
                        caminhoAntigo = arquivosSeparados[periodo]['caminho']
                        shutil.move(caminhoAntigo, os.path.join(pastaBackup, os.path.basename(caminhoAntigo)))
                        
                        arquivosSeparados[periodo] = {'data_hora': dataHora, 'caminho': os.path.join(pastaDestinoEcf, arquivo)}
                    else:
                        shutil.move(os.path.join(pastaDestinoEcf, arquivo), os.path.join(pastaBackup, arquivo))
                else:
                    arquivosSeparados[periodo] = {'data_hora': dataHora, 'caminho': os.path.join(pastaDestinoEcf, arquivo)}


def pesquisa(sped, dataInicio, dataFim, alinha, nome_empresa, cnpj):
        
    def icms():
        buscar_e_clicar('Arquivo')
        pyautogui.hotkey('ctrl','e')
        procurar_imagem(r'robo_bx\prints\input_sped.png')
        procurar_imagem(r"C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\inputSpedIcms.png")

        procurar_imagem(r"C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\inputTipoArquivo.png")
        procurar_imagem(r"C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\inputEscriFiscal.png")
        time.sleep(0.5)
        buscar_e_clicar("Estabelecimento", horizontal=+126, ocorrencia=2)
        buscar_e_clicar("transmitido *", horizontal=+210)
        
        procurar_imagem(r"C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\inputDataIni.png", confidence=0.6, horizontal=300)
        pyautogui.write(dataInicio)
        procurar_imagem(r"C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\inputDataFin.png", horizontal=300)
        pyautogui.write(dataFim)
        time.sleep(0.5)
        pyautogui.press('enter')
        pyautogui.hotkey('ctrl','p')
        resultadoPesquisa = verificarSolicitacao(alinha, sped)

        return resultadoPesquisa
    
    def piscofins():
        buscar_e_clicar('Arquivo')
        pyautogui.hotkey('ctrl','e')

        procurar_imagem(r'robo_bx\prints\input_sped.png')
        procurar_imagem(r'robo_bx\prints\inputSpedContribuicoes.png')
        procurar_imagem(r'robo_bx\prints\inputTipoArquivo.png', horizontal=10)
        procurar_imagem(r'robo_bx\prints\inputEscrituracao.png')
        procurar_imagem(r'robo_bx\prints\inputTipoPesquisa.png')
        procurar_imagem(r'robo_bx\prints\inputPeriodoEscrituracao.png')

        procurar_imagem(r'robo_bx\prints\inputDataInicio.png')
        pyautogui.write(dataInicio)

        procurar_imagem(r'robo_bx\prints\inputDataFim.png')
        pyautogui.write(dataFim)

        time.sleep(0.5)
        pyautogui.press('enter')
        pyautogui.hotkey('ctrl','p')
        resultadoPesquisa = verificarSolicitacao(alinha, sped)

        return resultadoPesquisa

    def ecf():
        buscar_e_clicar('Arquivo')
        pyautogui.hotkey('ctrl','e')

        procurar_imagem(r'robo_bx\prints\input_sped.png')
        procurar_imagem(r'robo_bx\prints\inputSpedEcf.png')
        procurar_imagem(r'robo_bx\prints\inputTipoArquivo.png', horizontal=10)
        procurar_imagem(r'robo_bx\prints\inputEscrituracao.png')
        procurar_imagem(r'robo_bx\prints\inputTipoPesquisa.png')
        procurar_imagem(r'robo_bx\prints\inputPeriodoEcf.png')

        procurar_imagem(r'robo_bx\prints\inputDataInicio.png')
        pyautogui.write(dataInicio)

        procurar_imagem(r'robo_bx\prints\inputDataFim.png')
        pyautogui.write(dataFim)

        time.sleep(0.5)
        pyautogui.press('enter')
        pyautogui.hotkey('ctrl','p')
        resultadoPesquisa = verificarSolicitacao(alinha, sped)

        return resultadoPesquisa

    def ecd():
        buscar_e_clicar('Arquivo')
        pyautogui.hotkey('ctrl','e')

        procurar_imagem(r'robo_bx\prints\input_sped.png')         
        procurar_imagem(r'robo_bx\prints\inputSpedEcd.png')
        procurar_imagem(r'robo_bx\prints\inputTipoArquivo.png', horizontal=10)
        procurar_imagem(r'robo_bx\prints\inputEscriDigital.png')

        procurar_imagem(r'robo_bx\prints\inputDataInicio.png')
        pyautogui.write(dataInicio)

        procurar_imagem(r'robo_bx\prints\inputDataFim.png')
        pyautogui.write(dataFim)

        time.sleep(0.5)
        pyautogui.press('enter')
        pyautogui.hotkey('ctrl','p')
        time.sleep(0.3)

        resultadoPesquisa = verificarSolicitacao(alinha, sped)
        print(f"Resultado retornado por ecd: {resultadoPesquisa}")
        return resultadoPesquisa
    
    match sped:
        case 'icms':
            resultadoPesquisa = icms()
            match resultadoPesquisa:
                case 'Nenhum Arquivo':
                    print(f"Não foi possivel realizar a pequisa dos arquivos -> {sped} no período -> {dataInicio} a {dataFim} || Não há arquivos transmitidos!")
                    pyautogui.press('enter')
                case 'Sem procuração':
                    print(f"Não foi possivel realizar a pequisa dos arquivos -> {sped} no período -> {dataInicio} a {dataFim} || Não há procuração habilitada!")
                    pyautogui.press('enter')                   
                case 'Tem Procuração':
                    print(f"Pesquisa do arquivos -> {sped} do período -> {dataInicio} a {dataFim} || Realizada com sucesso!")
                    procurar_imagem(r'robo_bx\prints\verificarResultado.png')
                    procurar_imagem(r'robo_bx\prints\inputCnpjPesquisa.png')
                    procurar_imagem(r'robo_bx\prints\solicitarArquivos.png')
                    coletarPedido(nome_empresa, cnpj)

        case 'piscofins':
            resultadoPesquisa = piscofins()
            match resultadoPesquisa:
                case 'Nenhum Arquivo':
                    print(f"Não foi possivel realizar a pequisa dos arquivos -> {sped} no período -> {dataInicio} a {dataFim} || Não há arquivos transmitidos!")
                    pyautogui.press('enter')
                case 'Sem procuração':
                    print(f"Não foi possivel realizar a pequisa dos arquivos -> {sped} no período -> {dataInicio} a {dataFim} || Não há procuração habilitada!")
                    pyautogui.press('enter')                   
                case 'Tem Procuração':
                    print(f"Pesquisa do arquivos -> {sped} do período -> {dataInicio} a {dataFim} || Realizada com sucesso!")
                    procurar_imagem(r'robo_bx\prints\verificarResultado.png')
                    procurar_imagem(r'robo_bx\prints\inputCnpjPesquisa.png')
                    procurar_imagem(r'robo_bx\prints\solicitarArquivos.png')
                    coletarPedido(nome_empresa, cnpj)

        case 'ecd':
            resultadoPesquisa = ecd()
            match resultadoPesquisa:
                case 'Nenhum Arquivo':
                    print(f"Não foi possivel realizar a pequisa dos arquivos -> {sped} no período -> {dataInicio} a {dataFim} || Não há arquivos transmitidos!")
                    pyautogui.press('enter')
                case 'Sem procuração':
                    print(f"Não foi possivel realizar a pequisa dos arquivos -> {sped} no período -> {dataInicio} a {dataFim} || Não há procuração habilitada!")
                    pyautogui.press('enter')                   
                case 'Tem Procuração':
                    print(f"Pesquisa do arquivos -> {sped} do período -> {dataInicio} a {dataFim} || Realizada com sucesso!")
                    procurar_imagem(r'robo_bx\prints\verificarResultado.png')
                    procurar_imagem(r'robo_bx\prints\inputCnpjPesquisa.png')
                    procurar_imagem(r'robo_bx\prints\situacaoSped.png'), time.sleep(1)
                    procurar_imagem(r'robo_bx\prints\situacaoSped.png')
                    buscar_e_clicar_todas('SUBSTITUÍDA', horizontal=-785, max_tentativas=5)
                    procurar_imagem(r'robo_bx\prints\solicitarArquivos.png')
                    coletarPedido(nome_empresa, cnpj)

        case 'ecf':
            resultadoPesquisa = ecf()
            match resultadoPesquisa:
                case 'Nenhum Arquivo':
                    print(f"Não foi possivel realizar a pequisa dos arquivos -> {sped} no período -> {dataInicio} a {dataFim} || Não há arquivos transmitidos!")
                    pyautogui.press('enter')
                case 'Sem procuração':
                    print(f"Não foi possivel realizar a pequisa dos arquivos -> {sped} no período -> {dataInicio} a {dataFim} || Não há procuração habilitada!")
                    pyautogui.press('enter')                   
                case 'Tem Procuração':
                    print(f"Pesquisa do arquivos -> {sped} do período -> {dataInicio} a {dataFim} || Realizada com sucesso!") 
                    procurar_imagem(r'robo_bx\prints\verificarResultado.png')
                    procurar_imagem(r'robo_bx\prints\inputContribPesquisa.png')
                    procurar_imagem(r'robo_bx\prints\solicitarArquivos.png')
                    coletarPedido(nome_empresa, cnpj)

def verificarSolicitacao(alinha, sped):
    max_espera = 10  
    inicio_espera = time.time()
    
    print("Aguardando a janela 'Pesquisando por arquivos...' abrir...")

    while time.time() - inicio_espera < max_espera:
        titulos = gw.getAllTitles()
        
        if 'Pesquisando por arquivos...' in titulos:
            print("Janela detectada! Prosseguindo...")
            break  
        
        time.sleep(0.5)  

    else:
        print("Janela não detectada no tempo limite.")
        return 'Erro: Janela não apareceu'  

    time.sleep(2)

    while True:
        print("Verificando janelas abertas...")
        titulos = gw.getAllTitles()

        contagemReceitanetBX = titulos.count('Receitanet BX')

        if contagemReceitanetBX > 1:
            print(f"Há {contagemReceitanetBX} janelas 'Receitanet BX' abertas.")
            print('Iniciando verificação de resultado...')

            max_tentativas = 5
            tentativas = 0
            resultadoPesquisa = None

            while tentativas < max_tentativas:
                print('Verificando')

                encontrou_sem_proc = procurar_imagem(r'robo_bx\prints\semProc.png', max_tentativas=3)

                if encontrou_sem_proc:
                    resultadoProc = 'semProc'
                    break

                encontrou_sem_proc2 = procurar_imagem(r'robo_bx\prints\procCancelada.png', max_tentativas=3)

                if encontrou_sem_proc2:
                    resultadoProc = 'procCancelada'
                    break

                encontrou_nenhum_arquivo = procurar_imagem(r'robo_bx\prints\nenhumArquivo.png', max_tentativas=3)

                if encontrou_nenhum_arquivo:
                    resultadoProc = 'nenhumArquivo'
                    break
                
                op_nao_concluida = buscar_e_clicar('OPERAÇÃO', max_tentativas=2)

                if op_nao_concluida:
                    resultadoProc = 'OPERAÇÃO'
                    break

                tentativas += 1
                time.sleep(1)
            
            match resultadoProc:
                case 'semProc':
                    resultadoPesquisa = 'Sem procuração'
                    pagiCaminhos.cell(column=3, row=alinha, value='SEM PROC')
                    planilha_caminhos.save(r'robo_bx\BasesNovas.xlsx')
                    print('Sem Procuração Salvo na Planilha')

                case 'procCancelada':
                    resultadoPesquisa = 'Sem procuração'
                    pagiCaminhos.cell(column=3, row=alinha, value='SEM PROC')
                    planilha_caminhos.save(r'robo_bx\BasesNovas.xlsx')
                    print('Sem Procuração Salvo na Planilha')

                case 'nenhumArquivo': 
                    resultadoPesquisa = 'Nenhum Arquivo'
                    pagiCaminhos.cell(column=3, row=alinha, value='SEM PROC')
                    planilha_caminhos.save(r'robo_bx\BasesNovas.xlsx')

                case 'OPERAÇÃO':
                    print('Erro no certificado, resolvendo e iniciando pesquisa novamente')
                    pyautogui.press('enter')
                    pyautogui.hotkey('ctrl','p')
                    resultadoPesquisa = verificarSolicitacao(alinha, sped)
                

            if resultadoPesquisa:
                print("Resultado Pesquisa:", resultadoPesquisa)
                return resultadoPesquisa  

        else:

            valor_atual = pagiCaminhos.cell(column=5, row=alinha).value

            if valor_atual:
                valores = valor_atual.split(", ")  
                if sped not in valores: 
                    valores.append(sped)
                    novo_valor = ", ".join(valores)
                else:
                    novo_valor = valor_atual  
            else:
                novo_valor = sped  

            resultadoPesquisa = 'Tem Procuração'
            print("Resultado Pesquisa:", resultadoPesquisa)
            pagiCaminhos.cell(column=5, row=alinha, value=novo_valor)
            planilha_caminhos.save(r'robo_bx\BasesNovas.xlsx')
            return resultadoPesquisa  

        time.sleep(1)


def coletarPedido(nome_empresa, cnpj):
    x, y, largura, altura = 918, 489, 70, 30  
    time.sleep(2)
    titulos = gw.getAllTitles()

    contagem = titulos.count('Receitanet BX')

    if contagem > 1:
        print(f"Há {contagem} janelas 'Receitanet BX' abertas.")
        printScr = pyautogui.screenshot(region=(917,492,72,22))
        printScr.save(r"robo_bx\prints\pedido.png")
        img = Image.open(r"robo_bx\prints\pedido.png")
        pedido = pytesseract.image_to_string(img, lang='por')
        pedido = pedido.strip()

        if pedido == '':
            print("pedido não esta vazio")

            while True:
                time.sleep(5.5)
                printScr = pyautogui.screenshot(region=(917,495,72,22))
                printScr.save(r"robo_bx\prints\pedido.png")
                img = Image.open(r"robo_bx\prints\pedido.png")
                pedido = pytesseract.image_to_string(img, lang='por')
                pedido = pedido.strip()
                if pedido != '':
                    break


        if pedido.isdigit():
            pedido_formatado = f'{int(pedido):,}'.replace(',', '.')
            print(f"Pedido formatado: {pedido_formatado}")
        else:
            print(f"Texto do pedido: {pedido}")

        procurar_imagem(r"robo_bx\prints\pedido.png")
        pyautogui.press('enter')
        downloadSped(pedido_formatado, nome_empresa, cnpj)

def downloadSped(pedido, nome_empresa, cnpj):
    buscar_e_clicar('Arquivo')
    pyautogui.hotkey('ctrl','a')
    time.sleep(1.5)
    procurar_imagem(r'C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\setaBaixo.png')
    time.sleep(1)

    """if not buscar_e_clicar(pedido):  
        time.sleep(2)  """

    pyautogui.hotkey('ctrl', 'u')  
    time.sleep(1)

    while True:
        encontrou = buscar_e_clicar(pedido, max_tentativas=15)
        
        if encontrou:
            print("Pedido Encontrado")
            break 
        
        print("Pedido nao encontrado em 25 tentativas, atualizando pág")
        pyautogui.hotkey('ctrl', 'u')
        pyautogui.sleep(2)


    procurar_imagem(r'C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\carregamentoDownload.png', clicks=0)
    time.sleep(2)
    procurar_imagem(r'C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\selecionarTodos.png', max_tentativas=500)
    procurar_imagem(r'C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\btnBaixar.png')

    print(f'Iniciando download da empresa: {cnpj} - {nome_empresa}')

    while True:
        if buscar_e_clicar("Não há arquivos na fila de download.", max_tentativas=2) is True:
            break

        else:
            time.sleep(10.5)

    procurar_imagem(r"C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\pedidosArquivos.png")
    procurar_imagem(r'C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\pedidoRealizado.png')
        
def fecharBx():
    procurar_imagem(r'robo_bx\prints\btnFechar.png')

def redirecionarSped(sped):
    download_dir = Path(r"C:\Users\gabriel.alvise\Documents\Arquivos ReceitanetBX")

    sped_nomes = {
        'ecd': 'ECD',
        'piscofins': 'PC',
        'icms': 'ICMS',
        'ecf': 'ECF'
    }

    def limpar_nome_pasta(nome):
        return re.sub(r'[<>:"/\\|?*]', '_', nome).strip().upper()

    def extrair_cnpj(arquivo, indice_cnpj):
        with open(arquivo, "r", encoding="latin-1") as f:
            primeira_linha = f.readline().strip()
            partes = primeira_linha.split("|")
            if len(partes) > indice_cnpj:
                cnpj_completo = re.sub(r'\D', '', partes[indice_cnpj])  
                if len(cnpj_completo) >= 12:  
                    cnpj_modificado = cnpj_completo[:11] + '1' + cnpj_completo[12:]  
                    return cnpj_modificado
                return cnpj_completo 
        return None

    def obter_nome_empresa(arquivo, indice_nome):
        with open(arquivo, "r", encoding="latin-1") as f:
            primeira_linha = f.readline().strip()
            partes = primeira_linha.split("|")
            if len(partes) > indice_nome:
                return limpar_nome_pasta(partes[indice_nome])
        return None

    def criar_pasta(cnpj, nome_empresa, sped):
        nome_empresa = limpar_nome_pasta(nome_empresa)
        pasta_empresa = download_dir / f"{cnpj} - {nome_empresa}"
        pasta_empresa.mkdir(parents=True, exist_ok=True)

        nome_sped = sped_nomes.get(sped, sped.upper())
        nome_sped = limpar_nome_pasta(nome_sped)
        pasta_sped = pasta_empresa / nome_sped
        pasta_sped.mkdir(parents=True, exist_ok=True)

        return pasta_sped

    def mover_arquivo(arquivo, destino):
        destino_arquivo = destino / arquivo.name
        shutil.move(str(arquivo), str(destino_arquivo))

    arquivos_txt = list(download_dir.glob("*.txt"))
    if not arquivos_txt:
        print("Erro: Nenhum arquivo .txt encontrado na pasta de downloads.")
        return

    indice_cnpj = {
        'ecd': 6,
        'piscofins': 9,
        'icms': 7,
        'ecf': 4
    }.get(sped, 4)

    indice_nome = {
        'ecd': 5,
        'piscofins': 8,
        'icms': 6,
        'ecf': 5
    }.get(sped, 5)

    empresas = {}  
    for arquivo in arquivos_txt:
        cnpj = extrair_cnpj(arquivo, indice_cnpj)
        if not cnpj:
            print(f"Erro: Não foi possível extrair o CNPJ do arquivo {arquivo.name}.")
            continue

        if cnpj not in empresas:
            nome_empresa = obter_nome_empresa(arquivo, indice_nome)
            empresas[cnpj] = nome_empresa if nome_empresa else "EMPRESA_DESCONHECIDA"

        pasta_empresa = criar_pasta(cnpj, empresas[cnpj], sped)
        mover_arquivo(arquivo, pasta_empresa)
        print(f"Arquivo '{arquivo.name}' movido para '{pasta_empresa}'")

        if sped == 'ecf':
            tratarSpedEcf(pasta_empresa, empresas[cnpj])

def dataSped(sped):
    match sped:
        case 'icms':
            dataInicio = '01/01/2020'
            dataFim = datetime.today().strftime('%d/%m/%Y')
        case 'piscofins':
            dataInicio = '01/01/2020'
            dataFim = datetime.today().strftime('%d/%m/%Y')
        case 'ecd':
            dataInicio = '01/01/2014'
            dataFim = datetime.today().strftime('%d/%m/%Y')
        case 'ecf':
            dataInicio = '01/01/2014'
            dataFim = datetime.today().strftime('%d/%m/%Y')

    return dataInicio, dataFim

def extracaoCnpj(alinha):
    cnpjs = []
    for row in pagiCaminhos.iter_rows(min_row=alinha):  
        nome_empresa = row[0].value  
        cnpj = row[1].value 
        if cnpj:
            cnpjs.append((nome_empresa, cnpj))  

    return cnpjs

def abrirBx():
    alinha = 100
    sped = 'icms'
    cnpjs = extracaoCnpj(alinha)  
    dataInicio, dataFim = dataSped(sped)

    with tqdm(total=len(cnpjs), desc="Processando CNPJs", ncols=70, unit="CNPJ") as barra:
        for nome_empresa, cnpj in cnpjs:  

            print(f"Processando CNPJ: {cnpj} - {nome_empresa}")  

            start_time = time.time()  
            procurar_imagem(r'C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\bx_areadetrabalho.png', clicks=2)

            while procurar_imagem(r"C:\Users\gabriel.alvise\Desktop\ROBOS\robo_bx\prints\bxcarregando.png", clicks=0, max_tentativas=2):
                time.sleep(1)

            set_high_priority("javaw.exe")

            login(cnpj)
            pesquisa(sped, dataInicio, dataFim, alinha, nome_empresa, cnpj)
            time.sleep(1)
            redirecionarSped(sped)

            time.sleep(1.2)
            fecharBx()

            alinha += 1  

            barra.update(1)  
            tempo_gasto = time.time() - start_time
            tempo_restante = (len(cnpjs) - barra.n) * tempo_gasto
            print(f"Tempo estimado para conclusão: {tempo_restante:.2f} segundos")
    

abrirBx()

#pesquisa('ecd', '01/01/2014', '31/12/2024')

#titulos = gw.getAllTitles()
#print(titulos)
